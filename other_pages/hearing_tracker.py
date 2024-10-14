import os
import numpy as np
import streamlit as st
import pandas as pd
import json
import datetime
import pytz
import re
from urllib.parse import quote_plus

from custom_module.mega.mega.mega import Mega
import gdown

from other_pages.googleapi import download_data,download_sheet,upload_data
from openpyxl.utils import get_column_letter
from streamlit.components.v1 import html as display_html

# for getting cover image
import eyed3
from io import BytesIO
from PIL import Image

class SP_hearing_Class:
    def __init__(self):
        
        self.page_config = {'page_title': "Shravanam",
                            'page_icon':'💊',
                            'layout':'centered'}
        self.page_map = {
            'SP':self.sp_lectures,
            'SP_lec_player':self.hear_sp_now,
            "user_registration":self.registration
            }
        self.current_page = 'SP'
        self.mega = Mega()
        
        # ===================================databases=================================
        # for Prabhupada lectures------------------------------------------------------
        self.sp_sindhu_df = pd.read_csv("./local_data/SP_sindhu_config.csv")
        self.sp_sindhu_df.insert(0,"select",False)
        
        # creating the lec-url
        rootURL = 'http://localhost:8501/'
        # rootURL = 'http://localhost:8501/'        
        self.sp_sindhu_df['lec_url'] = rootURL+'?target=hear-now&source=sp_sindhu&id='\
            + self.sp_sindhu_df['encrypt_id']
        
        self.sp_sindhu_df['name'] = np.where(self.sp_sindhu_df['category'].isin(['Bg','SB']),
                                             self.sp_sindhu_df['prefix'] + " "+self.sp_sindhu_df['name'],
                                             self.sp_sindhu_df['name'])
        # user data (this is used if logged in)
        self._sp_userdb = None
        self._sp_userdb_refresh = True
        
        # prabhupada hearing page db
        self._sp_single_choice_dfdict = {
                                    # global
                                     'update_post_date':True,
                                    
                                    # for SB related
                                     'SB_canto':None,
                                     'SB_chapter':None,
                                     'SB_location':None,
                                     
                                     'update_post_canto':True,
                                     'update_post_chapter':True,
                                     
                                     # for BG related
                                     'update_post_bg_chapter':True,
                                     'BG_chapter':None,
                                     'BG_location':None,
                                     
                                     # for all other
                                     'update_post_radio':True,
                                     'update_post_year':True,
                                     'update_post_month':True,
                                     'yeardf':None,
                                     'monthdf':None,
                                     'locationdf':None
                                     }
        self.update_sp_sindhu_finder("category == 'SB'",['canto','chapter','location'])
        self.update_sp_sindhu_finder("category == 'Bg' ",['BG_chapter','BG_location'])
        
        # for hear now
        self.play_now_info_dict = None
        # {
        #     'encrypt_id':None,
        #     'mega_id':None,
        #     'sp_id':None,
        #     'lecture_name':None,
        # }
        
        # for user info etc-------------------------------
        self._userdb = None
        self.refresh_userdb = True
        
        # userdata from sp-sindhu table
        self.userinfo = {"mode":"guest",
                         "user":{}
                         }
        
        
        self.registrationinfo = {"source":'external',# bdv,external
                                 'reg_status':None, # success after submission
                                }
    def perform_login(self,username,password,callmode):
        """
        * perform following if callmode=submit
        * update self.userinfo
        * update self._userdb['dfself']
        * update self.sp_sindhu_df
        * if callmode = ask
        * returns 1 if valid username and password
        """
        _user_is_valid = 0
        _password_is_correct = 0
        _userinfo = None
        if username in self.userdb['dict']['existing_userid_list']:
            _user_is_valid=1
            
            _userinfo = json.loads(
                    self.userdb['dfall']\
                    .query("spid=='userinfo'")[username].tolist()[0])
            _user_dbcol = json.loads(
                self.userdb['dfall']\
                    .query("spid=='dbcol'")[username].tolist()[0])
            
            if password==_userinfo['creds']['password']:
                _password_is_correct = 1
        
        if callmode=='ask':
            # 0 if no valid user
            # 1 if user is valid and password is incorrect
            # 2 if valid user and correct password
            return _user_is_valid + _password_is_correct
        
        elif callmode =='submit':
            if _user_is_valid + _password_is_correct ==2:
                # do all the tasks
                
                # initialize the userdb
                _ = self.userdb
                
                # update self.userinfo
                self.userinfo = {"mode":'user',
                                 'user':_userinfo,
                                 'dbcol':_user_dbcol}
                # update self._userdb['dfself']
                _columns = ['dbrow','spid',username]
                self._userdb['dfself'] = self._userdb['dfall'][_columns].copy(deep=True)
                
                # update self.sp_sindhu_df
                # convert id to numbers
                dfself = self.userdb['dfself'].copy(deep=True)
                dfself['spid'] = pd.to_numeric(dfself['spid'],errors='coerce')
                dfself.dropna(subset='spid',inplace=True)
                self.sp_sindhu_df = self.sp_sindhu_df.merge(dfself,left_on='id',
                                        right_on='spid',
                                        how='left',indicator=True)
                # convert the lecture_status to dict                
                # print(self.sp_sindhu_df[[username]].head()[username].tolist())
                self.sp_sindhu_df[username] = self.sp_sindhu_df[username]\
                    .apply(lambda x: {"status":'pending'} if x=='' else json.loads(x))
                # print("while loggin on")
                # print(self.sp_sindhu_df.columns)
                
                # update the dicts
                self._refresh_single_choice_df_dict()
                st.snow()
    
    def _refresh_single_choice_df_dict(self):
        
        self.update_sp_sindhu_finder("category == 'SB'",['canto','chapter','location'])
        self.update_sp_sindhu_finder("category == 'Bg' ",['BG_chapter','BG_location'])
        
        self._sp_single_choice_dfdict['update_post_radio'] = True
        self._sp_single_choice_dfdict['update_post_date'] = True
        
    
    @property
    def user_exists(self):
        if hasattr(self,'userinfo'):
            return self.userinfo['mode'] == 'user'
        else:
            return False
        
    @property
    def userdb(self):
        """
        dfself, dfall, dict
        """
        if self.refresh_userdb:
            dbarray = download_sheet(1,'sp_sindhu_creds')
            dbdf = pd.DataFrame(dbarray[1:],columns=dbarray[0])
            
            dbdict = {k:v for k,v in dict(zip(dbdf.mdata_key,
                              dbdf.mdata_value)).items() if k not in ['3','']}
            dbdict['existing_userid_list'] = dbdict['existing_userid_list'].split(",")
            
            self._userdb = {
                'dfself':None,
                'dfall':dbdf.drop(columns=['mdata_key','mdata_value']),
                'dict':dbdict
                }
            self.refresh_userdb = False
            
        return self._userdb
    
    @property
    def bdvapp(self):
        return st.session_state['bdv_app']
    
# ================functions related to Srila Prabhupada Sindhu hearing
    def get_sp_sindhu_value_count(self,query,column_name,new_name):
        df = self.sp_sindhu_df.copy(deep=True)
        dfsummary = df.query(query)[column_name]\
        .value_counts().reset_index()
        dfsummary.columns = [new_name,'count']
        final_col_type = 'str' if new_name in ['month','location'] else 'int'
        dfsummary[new_name] = dfsummary[new_name].astype(final_col_type)
        
        if self.user_exists:
            user_phn = str(self.userinfo['user']['creds']['phone'])
            df2 = df[df[user_phn].apply(lambda x: x['status']=='completed')]
            dfstatus_summary = df2.query(query)[column_name]\
            .value_counts().reset_index()
            dfstatus_summary.columns= [new_name,'done_count']
            
            # now merge with original
            dfsummary = dfsummary.merge(dfstatus_summary,on=new_name,how='left')
            dfsummary['done_count'] = dfsummary['done_count'].fillna(0)
            dfsummary['done_count'] = dfsummary['done_count'].astype('int')
            # st.dataframe(dfsummary)
            
            dfsummary['overall_count'] = dfsummary['done_count'].astype('str')\
                                    +'/'+ dfsummary['count'].astype('str')
            # st.dataframe(dfsummary)
            
            dfsummary = dfsummary[[new_name,'overall_count']]
            dfsummary.columns = [new_name,'count']
        
        # sort the values
        if new_name =='month':
            monthmap = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
            dfsummary['newcol'] = dfsummary['month'].apply(lambda x:monthmap[x])
            dfsummary = dfsummary.sort_values(by='newcol',ascending=True)
            dfsummary.drop(columns='newcol',inplace=True)
        else:
            dfsummary = dfsummary.sort_values(by=new_name,ascending=True)
        
        dfsummary.insert(0,'select',False)
        return dfsummary.copy()
    
    def update_sp_sindhu_finder(self,query,update_list):
        """
        #### for SB
        * canto
        * chapter
        * location
        #### for BG
        * BG_chapter
        * BG_location
        
        #### for other
        * yeardf
        * monthdf
        * locationdf
        """
        # for SB
        if 'canto' in update_list:
            self._sp_single_choice_dfdict['SB_canto'] = self.get_sp_sindhu_value_count(query,'sb_canto','canto')
        
        if 'chapter' in update_list:
            self._sp_single_choice_dfdict['SB_chapter'] = self.get_sp_sindhu_value_count(query,'sb_ch','chapter')
        
        if 'location' in update_list:
            self._sp_single_choice_dfdict['SB_location'] = self.get_sp_sindhu_value_count(query,'location','location')
        
        # for BG
        if 'BG_chapter' in update_list:
            self._sp_single_choice_dfdict['BG_chapter'] = self.get_sp_sindhu_value_count(query,'bg_ch','chapter')
        
        if 'BG_location' in update_list:
            self._sp_single_choice_dfdict['BG_location'] = self.get_sp_sindhu_value_count(query,'location','location')
        
        # for all other
        if 'yeardf' in update_list:
            self._sp_single_choice_dfdict['yeardf'] = self.get_sp_sindhu_value_count(query,'year','year')
        
        if 'monthdf' in update_list:
            self._sp_single_choice_dfdict['monthdf'] = self.get_sp_sindhu_value_count(query,'month_Eng','month')
        
        if 'locationdf' in update_list:
            self._sp_single_choice_dfdict['locationdf'] = self.get_sp_sindhu_value_count(query,'location','location')
            
    def __single_select_chb_sp(self,editor_key,original_key,checkbox_column):
        """
        * editor_key = key for the st.data_editor
        * original_key = the df should be stored in st.session_state[original_key]
        * checkbox_column = name of the checkbox column
        * It must be the first column"""
        edited_rows = st.session_state[editor_key]['edited_rows']
        if len(edited_rows)==1:
            row_num = list(edited_rows.keys())[0]
            # df = st.session_state[original_key].copy(deep=True)
            df = self._sp_single_choice_dfdict[original_key].copy()
            
            set_2_True = list(edited_rows.values())[0][checkbox_column]        
            if set_2_True:
                df[checkbox_column] = False
                df.iloc[row_num,0] = True
                self._sp_single_choice_dfdict[original_key]= df.copy()
            else:
                df[checkbox_column] = False
                self._sp_single_choice_dfdict[original_key]= df.copy()
                st.session_state[editor_key]['edited_rows'] = {}
        
        # for SB
        if editor_key == 'sb_canto_selector':
            self._sp_single_choice_dfdict['update_post_canto'] = True
        
        elif editor_key == 'sb_chapter_selector':
            self._sp_single_choice_dfdict['update_post_chapter'] = True        
        # for bG
        elif editor_key == 'bg_chapter_picker':
            self._sp_single_choice_dfdict['update_post_bg_chapter'] = True
        # for others
        elif editor_key == 'yeardf_picker':
            self._sp_single_choice_dfdict['update_post_year'] = True
        
        elif editor_key == 'monthdf_picker':
            self._sp_single_choice_dfdict['update_post_month'] = True
            
    def sp_lectures(self):
        
        if self.bdvapp.userinfo:
            if 'sp_vani_admin' in self.bdvapp.userinfo['roles']:
                st.button("Pending Registrations",on_click=self.switch_page,
                        args=['user_registration'])
        
        st.header(":rainbow[Srila Prabhupada Ki Jai!!]")
        if self.user_exists:
            st.markdown(f"### :gray[Hare Krishna] :rainbow[{self.userinfo['user']['creds']['name']}]")
        
        spdf = self.sp_sindhu_df.copy(deep=True) # load the lecture config file
        # displayconfig = {
        #     'select':st.column_config.CheckboxColumn('pick'),
        #     'category':st.column_config.TextColumn("type",width='small'),
        #     # 'prefix':st.column_config.TextColumn("id"),
        #     'name':st.column_config.TextColumn("title"),
        #     # 'name_1':st.column_config.TextColumn("title"),
        #     'url':st.column_config.LinkColumn("url",display_text="download"),
        #     'month_Eng':st.column_config.TextColumn("month"),
        #     'mmm_dd':st.column_config.TextColumn("day"),
        #     'yy_mmm_dd':st.column_config.TextColumn("date"),
        #     'lec_url':st.column_config.LinkColumn("link",display_text='hear'),
        #     'duration':st.column_config.NumberColumn("⏳min")
        # }
        
        default_count_config = {"count":\
                                st.column_config.TextColumn("#",width='small',
                                                            help='heard/total',disabled=True) if self.user_exists else\
                                st.column_config.NumberColumn("#",width='small',
                                                            help='Total Count',disabled=True),
                                'select':st.column_config.CheckboxColumn("☑",help='Check to filter',disabled=False)
                                }
        filter_config = {'canto':{
            **default_count_config,
            'canto':st.column_config.NumberColumn('c',help='canto',width='small',disabled=True)
            },
                        'chapter':{
            **default_count_config,
            'chapter':st.column_config.NumberColumn("ch",help='chapter',width='small',disabled=True)
            },
                        'location':{
            **default_count_config,
            'location':st.column_config.TextColumn("loc",help='Location',width='small',disabled=True)
            },
                        'year':{
            **default_count_config,
            'year':st.column_config.NumberColumn("yy",help='year',width='small',disabled=True,format='%d')
            },
                        'month':{
            **default_count_config,
            'month':st.column_config.TextColumn("mon",help='month',width='small',disabled=True)
            },
                        }
        # st.dataframe(spdf,column_config=displayconfig)
        
        subsubpage_dict = {1:'SB',
                           2:'BG',
                           3:'MW, RC etc'
                           }
        def sectionradiohandler():
            self._sp_single_choice_dfdict['update_post_date'] = True
            
        section_index = st.radio("Choose a Section",
                              options=[1,2,3],
                              index=0,
                              format_func=lambda x: subsubpage_dict[x],
                              horizontal=True,
                              on_change=sectionradiohandler,
                              key='sb_bg_other_radio'
                              )
        
        if section_index==1:
            querysofar = ["category == 'SB'"]
        
        elif section_index ==2:
            querysofar = ["category == 'Bg' "]
            
        elif section_index == 3:
            def update_on_change_radio():
                self._sp_single_choice_dfdict['update_post_radio'] = True
                self._sp_single_choice_dfdict['update_post_date'] = True
                
            further_type = st.radio("Choose further",
                     options=['MW','RC',"Interv",'Lect','Fest','Cc',"Addr",'Story'],
                     on_change=update_on_change_radio,
                     format_func=lambda x:{"MW":'Morning Walks',
                                           'RC':'Room Conversation',
                                           'Lect':"Public Lectures",
                                           'Cc':"CC Lectures",
                                           'Fest':"Festivals",
                                           "Interv":"Interviews",
                                           "Addr":"Addresses",
                                           'Story':"Stories"
                                           }[x])
            querysofar = [f"category =='{further_type}'"]
            if further_type == 'Addr':
                querysofar = [f"((category == 'Arr') or  (category == 'Addr')) "]
            
            if self._sp_single_choice_dfdict['update_post_radio']:
                self.update_sp_sindhu_finder(" and ".join(querysofar),
                                             ['yeardf','monthdf','locationdf'])
                self._sp_single_choice_dfdict['update_post_radio'] = False
        
        # this filters based on section selection
        spdf = spdf.query(" and ".join(querysofar))
        
        # -------------provision for date filter
        def checkbox_handler():
            # if not st.session_state['date_filter_checkbox']:
            if section_index==1:
                self.update_sp_sindhu_finder("category == 'SB'",
                                            ['canto','chapter','location'])
            elif section_index == 2:
                self.update_sp_sindhu_finder("category == 'Bg' ",
                                                ['BG_chapter','BG_location'])
            elif section_index ==3:
                self.update_sp_sindhu_finder(querysofar[0],
                                                ['yeardf','monthdf','locationdf'])
        
        if section_index==3 and further_type=="Story":
            # no date ilter for stories
            pass
        
        elif st.checkbox("Add date filter",
                        key='date_filter_checkbox',
                        on_change=checkbox_handler):
            def sp_date_handler():
                self._sp_single_choice_dfdict['update_post_date'] = True
            
            chosen_date = st.date_input("Enter date", 
                                        format="YYYY-MM-DD",
                                        min_value=datetime.date(1966,2,19), # hardcoded from excel
                                        max_value=datetime.date(1977,7,1), # hardcoded from excel
                                        value=datetime.date(1972,
                                                            datetime.datetime.today().month,
                                                            datetime.datetime.today().day),
                                        on_change=sp_date_handler)
            chosen_year = chosen_date.year
            chosen_month = chosen_date.month
            chosen_day = chosen_date.day
            
            querysofar.extend([f"year == {chosen_year}", f"month == {chosen_month}"])
            if not st.toggle("Ignore day", value=True,on_change=sp_date_handler):
                querysofar.append(f"day == {chosen_day}")
            
            spdf = spdf.query(" and ".join(querysofar))
            if self._sp_single_choice_dfdict['update_post_date']:
                if section_index==1:
                    self.update_sp_sindhu_finder(' and '.join(querysofar),
                                                ['canto','chapter','location'])
                elif section_index ==2:
                    self.update_sp_sindhu_finder(' and '.join(querysofar),
                                                 ['BG_chapter','BG_location'])
                elif section_index==3:
                    self.update_sp_sindhu_finder(" and ".join(querysofar),
                                                 ['yeardf','monthdf','locationdf'])
                self._sp_single_choice_dfdict['update_post_date'] = False
        # date filter ends here
        
        if section_index == 1:
            # further filters of canto- chapter location
            st.divider()
            left,middle,right = st.columns(3)
            with left:
                st.markdown("Choose Canto")
                list_chosen_canto = st.data_editor(
                    self._sp_single_choice_dfdict['SB_canto'],
                    hide_index=True,
                    column_config=filter_config['canto'],
                    key='sb_canto_selector',
                    on_change=self.__single_select_chb_sp,
                    args=['sb_canto_selector','SB_canto','select']
                    ).query("select == True")['canto'].tolist()
                
                if list_chosen_canto :
                    querysofar.append(f"sb_canto == {list_chosen_canto[0]}")
                    spdf = spdf.query(" and ".join(querysofar) )
                    
                # update only when canto is changed
                if self._sp_single_choice_dfdict['update_post_canto']:
                    self.update_sp_sindhu_finder(' and '.join(querysofar),
                                                ['chapter','location'])
                    self._sp_single_choice_dfdict['update_post_canto'] = False        
            
            with middle:
                st.markdown("Choose Chapter")
                list_chosen_chapter = st.data_editor(
                    self._sp_single_choice_dfdict['SB_chapter'],
                    hide_index=True,
                    column_config=filter_config['chapter'],
                    key='sb_chapter_selector',
                    on_change=self.__single_select_chb_sp,
                    args=['sb_chapter_selector','SB_chapter','select']
                    ).query("select == True")['chapter'].tolist()
                
                if list_chosen_chapter:
                    querysofar.append(f" sb_ch == {list_chosen_chapter[0]}")
                    spdf = spdf.query(" and ".join(querysofar))
                
                # update only if chapter edited
                if self._sp_single_choice_dfdict['update_post_chapter']:
                    self.update_sp_sindhu_finder(' and '.join(querysofar),
                                                    update_list=['location'])
                    self._sp_single_choice_dfdict['update_post_chapter'] = False
                            
            with right:
                st.markdown("Choose Location")
                list_chosen_location = st.data_editor(
                    self._sp_single_choice_dfdict['SB_location'],
                    hide_index=True,
                    column_config=filter_config['location'],
                    key='sb_location_selector',
                    on_change=self.__single_select_chb_sp,
                    args=['sb_location_selector','SB_location','select']
                    ).query("select == True")['location'].tolist()
                
                if list_chosen_location:
                    spdf = spdf.query(f"location == '{list_chosen_location[0]}' ")            
        
        elif section_index==2:
            # BG
            st.divider()
            left,right = st.columns(2)
            with left:
                st.markdown("choose chapter")
                list_chosen_chapter = st.data_editor(
                    data=self._sp_single_choice_dfdict['BG_chapter'],
                    hide_index=True,
                    column_config=filter_config['chapter'],
                    key='bg_chapter_picker',
                    on_change=self.__single_select_chb_sp,
                    args=['bg_chapter_picker','BG_chapter','select']
                ).query("select == True")['chapter'].tolist()
                
                if list_chosen_chapter:
                    querysofar.append(f"bg_ch == {list_chosen_chapter[0]} ")
                    spdf = spdf.query(" and ".join(querysofar))
                
                if self._sp_single_choice_dfdict['update_post_bg_chapter']:
                    self.update_sp_sindhu_finder(" and ".join(querysofar),
                                                 ['BG_location'])
                    self._sp_single_choice_dfdict['update_post_bg_chapter'] = False
            
            with right:
                st.markdown("Choose location")
                list_chosen_location = st.data_editor(
                    data=self._sp_single_choice_dfdict['BG_location'],
                    hide_index=True,
                    column_config=filter_config['location'],
                    on_change=self.__single_select_chb_sp,
                    key='bg_location_picker',
                    args=['bg_location_picker','BG_location','select']
                ).query("select==True")['location'].tolist()
                
                if list_chosen_location:
                    spdf = spdf.query(f"location == '{list_chosen_location[0]}' ")            
                
        elif section_index == 3:
            # morning Waslk and Room Conversations
            if further_type!='Story':
                st.divider()
                left,middle,right = st.columns(3)
                with left:
                    st.markdown('Choose Year')
                    list_chosen_year = st.data_editor(
                        data=self._sp_single_choice_dfdict['yeardf'],
                        hide_index=True,
                        column_config=filter_config['year'],
                        on_change=self.__single_select_chb_sp,
                        key='yeardf_picker',
                        args=['yeardf_picker','yeardf','select']
                    ).query("select == True")['year'].tolist()

                    if list_chosen_year:
                        querysofar.append(f"year == {list_chosen_year[0]}")
                        spdf = spdf.query(" and ".join(querysofar))
                    
                    if self._sp_single_choice_dfdict['update_post_year']:
                        self.update_sp_sindhu_finder(" and ".join(querysofar),
                                                    ['monthdf','locationdf'])
                        self._sp_single_choice_dfdict['update_post_year'] = False
                
                with middle:
                    st.markdown("Choose month")
                    list_chosen_month = st.data_editor(
                        data=self._sp_single_choice_dfdict['monthdf'],
                        hide_index=True,
                        column_config=filter_config['month'],
                        on_change=self.__single_select_chb_sp,
                        key='monthdf_picker',
                        args=['monthdf_picker','monthdf','select']
                    ).query("select == True")['month'].tolist()
                    
                    if list_chosen_month:
                        querysofar.append(f"month_Eng == '{list_chosen_month[0]}'")
                        spdf = spdf.query(" and ".join(querysofar))
                    
                    if self._sp_single_choice_dfdict['update_post_month']:
                        self.update_sp_sindhu_finder(" and ".join(querysofar),
                                                    ['locationdf'])
                        self._sp_single_choice_dfdict['update_post_month'] = False
                
                with right:
                    st.markdown("Choose Location")
                    list_chosen_location = st.data_editor(
                        data = self._sp_single_choice_dfdict['locationdf'],
                        hide_index=True,
                        column_config=filter_config['location'],
                        key='locationdf_picker',
                        on_change=self.__single_select_chb_sp,
                        args=['locationdf_picker','locationdf','select']
                    ).query("select==True")['location'].tolist()

                    if list_chosen_location:
                        spdf = spdf.query(f"location == '{list_chosen_location[0]}' ")            
            #now display the list of filtered lectures
            # spdf
        
        # now we have all the filters applied. Just display the filtered data
        
        st.divider()        
        st.dataframe(spdf)
        
        search_query = st.text_input(":gray[Enter query]",max_chars=15)
        if search_query:
            spdf = spdf[spdf.name.str.lower().str.contains(search_query.strip().lower())]
        
        st.caption(f"found {len(spdf)} records")
        # this is table view. Links doesn't work well in this on mobile
        # st.data_editor(spdf,column_config=sb_column_config,
        #                 column_order=list(sb_column_config.keys()),
        #                 hide_index=True,
        #                 disabled=list(set(sb_column_config.keys())-{'select'})).query("select==True")
        if self.user_exists:
            user_phone_number = str(self.userinfo['user']['creds']['phone'])
        for _row, row in spdf.reset_index(drop=True).iterrows():
            lec_name = row['name']
            duration = row['duration']
            lec_url = row['lec_url']
            download_url = row['url']
            status_color = 'gray'
            if self.user_exists:
                lecture_status = row[user_phone_number]['status']
                status_color = {'pending':'red',
                                'in_progress':'orange',
                                'completed':'green'}[lecture_status]
            
            with st.expander(f":gray[{_row+1}\.] :{status_color}[{lec_name} ({duration})]"):
                st.markdown(f"[link to hear]({lec_url})")
            if (_row+1) %20==0:
                chb_placeholder= st.empty()
                chb_response = chb_placeholder.checkbox(f":blue[Show `{len(spdf)-1-_row}` more]",
                                    key=f'showmore_{_row}',
                                    value=False)
                if not chb_response:
                    break
                else:
                    chb_placeholder.empty()

    def hear_sp_now(self):
        url = f"https://mega.co.nz/#!{self.play_now_info_dict['mega_id']}"
        destination = './local_data/sp_storage'
        filename = f"{self.play_now_info_dict['sp_id']}.mp3"
        displayname = self.play_now_info_dict['lecture_name']
        
        available_files = os.listdir(destination)
        
        msgbox = st.empty()
        if filename not in available_files:
            MAX_LEC_STORE = 20
            # keep the latest MAX_LEC_STORE files and delete the older ones
            # reverse = True sorted in decending order
            available_files = sorted(available_files,
                                     key=lambda x: os.path.getmtime(f"{destination}/{x}"),reverse=True)
            for one_file in available_files[MAX_LEC_STORE:]:
                os.remove(f"{destination}/{one_file}")
                
            with msgbox.container():
                st.markdown(f"downloading from mega ")
                st.caption(displayname)
                st.caption("This may take a while please wait ...")
                self.mega.download_url(url,destination,dest_filename=filename)
                st.success("completed")
                
        msgbox.markdown(f"## :rainbow[{displayname}]")
        st.markdown("")
        st.markdown("")
        # display the image if the audio file have one
        eye_file = eyed3.load(f"{destination}/{filename}")
        file_duration_secs = eye_file.info.time_secs
        if eye_file.tag.images:
            cover_image = Image.open(BytesIO(eye_file.tag.images[0].image_data))
            st.image(cover_image)
            
        st.markdown("")
        st.markdown("")
        foward_min = st.number_input("forward (in min)",step=1,min_value=0)
        
        st.markdown("")
        st.markdown("")
        st.markdown("")
        st.audio(f"{destination}/{filename}",format="audio/wav",
                 start_time=foward_min*60)

        st.markdown(f"[Download this mega]({url})")
        
        if self.user_exists:
            st.divider()
            st.markdown(f"### :gray[Hare Krishna ] :rainbow[{self.userinfo['user']['creds']['name']}]")            
            
            hearing_status = self.sp_sindhu_df.query(f"spid == {self.play_now_info_dict['sp_id']}")\
            [str(self.userinfo['user']['creds']['phone'])].tolist()[0]
            if hearing_status['status'] == 'pending':
                st.error("You have not heard this lecutre")
            elif hearing_status['status'] == 'in_progress':
                last_modification_time = datetime.datetime.strptime(hearing_status['last_modification_time'],"%Y%b%d%a %H%M%S")\
                    .strftime("%d %b %Y @ %H:%M")
                st.warning("### You have heard this lecture until")
                st.markdown(f"#### `{hearing_status['heard_until']}` :gray[minutes]")
                st.markdown(f"#### :gray[last heard on] {last_modification_time}")
            else:
                # have heard this lecture
                last_modification_time = datetime.datetime.strptime(hearing_status['last_modification_time'],"%Y%b%d%a %H%M%S")\
                    .strftime("%d %b %Y @ %H:%M")
                st.success("### You have heard this lecture on ")
                st.markdown(f"#### {last_modification_time}")
                
            spid = int(self.play_now_info_dict['sp_id'])
            
            dbrow = self.sp_sindhu_df[['spid','dbrow']].query(f"spid == {spid}")['dbrow'].tolist()[0]
            dbcol = self.userinfo['dbcol']
            dbsheet = 'sp_sindhu_creds'
            
            india_timezone = pytz.timezone('Asia/Kolkata')
            timestamp = datetime.datetime.now(india_timezone).strftime("%Y%b%d%a %H%M%S")
            
            
            def update_lecture_status(status_dict,dbrow,dbcol,dbname):
                status_json = json.dumps(status_dict)
                upload_data(db_id=1,
                            range_name=f"{dbname}!{get_column_letter(dbcol)}{dbrow}",
                            value=[[status_json]])
                # update the local data
                user_phone = str(self.userinfo['user']['creds']['phone'])
                self.sp_sindhu_df[user_phone] = np.where(self.sp_sindhu_df['spid']==spid,
                                                         status_dict,
                                                         self.sp_sindhu_df['spid'])
                
            
            new_status = st.radio(":gray[Update Status of Lecture]",
                     options=['in progress','completed'],
                     index=1 if hearing_status['status']=='completed' else 0,
                     disabled= True if hearing_status['status']=='completed' else False)
            
            if new_status =='in progress':
                duration = st.number_input("Heard until (in minute)",min_value=0,step=1)
                
                if duration:
                    status_prog = {'status':'in_progress',
                                   'heard_until':duration,
                                   'lecture_notes':'',
                                   'last_modification_time':timestamp
                                   }
                    st.button("Update status",on_click=update_lecture_status,
                              args=[status_prog,dbrow,dbcol,dbsheet],
                              type='primary')
                    
            elif new_status =='completed':
                MIN_LINE = int(max(2,min(12,(5*file_duration_secs)/(30*60))))
                MIN_WORD = int(max(10,min(80,(50*file_duration_secs)/(30*60))))
                
                lec_notes = st.text_area("Lecture Summary",
                                         help=f"Must write at least {MIN_LINE} lines and minimum {MIN_WORD} words in order to mark as completed",
                                         value="" if hearing_status['status']!='completed' else hearing_status['lecture_notes'],
                                         max_chars=400)
                n_lines, n_words = len(lec_notes.splitlines()),len(lec_notes.split())
                st.caption(f"you have written {n_lines} lines {n_words} words")
                if n_lines>MIN_LINE-1 and n_words>MIN_WORD-1:
                                      
                    status_compl = {'status':'completed',
                                   'heard_until':-1,
                                   'lecture_notes':lec_notes,
                                   'last_modification_time':timestamp
                                   }
                    st.button("Update status",on_click=update_lecture_status,
                              args=[status_compl,dbrow,dbcol,dbsheet],
                              type='primary')
                    
                    
                else:
                    if n_lines <MIN_LINE:
                        st.caption(f"{MIN_LINE-n_lines} more lines are needed")
                    if n_words<MIN_WORD:
                        st.caption(f"{MIN_WORD-n_words} more words are needed")
                
            
            
            
            
        
        
    def switch_page(self,newpage):
        self.current_page=newpage
    
    def registration(self):
        if self.registrationinfo['reg_status'] == 'success':
            st.success("You have successfully submitted your form")
            st.info("Upon verification your account will be created")
            
        elif self.registrationinfo=='bdv':
            st.success("Congratulations!!")
            st.subheader(":rainbow[You are already part of BDV family]")
            st.markdown("### :gray[Just enter your phone number to register]")
            with st.popover("Why are we asking phone number"):
                st.markdown("1\. To protect against robotic attacks")
                st.markdown("2\. To have a unique id of user")
                st.markdown("2\. To connect with you if any problems 🤝")
            number = st.number_input("Please enter 10 digit phone number")
            if number:
                if len(str(number)) !=10:
                    st.error("Please enter 10 digit number")
                elif not all([i in '0123456789' for i in str(number)]):
                    st.error("Only Numbers are allowed")
                
                elif str(number) in self.userdb['dict']['existing_userid_list']:
                    st.error("A user already exists with this number")
                else:
                    # success
                    def register_user(number):
                        password = self.bdvapp.userinfo['password']
                        name = f"{self.bdvapp.userinfo['name']} Pr"
                        userinfo = {
                                    "account_is_active":'no',
                                    "creds":{'name':name,
                                             'password':password,
                                             'phone':number}
                                    }
                        id_insert_col = int(self.userdb['dict']['insert_id_col_index'])
                        upload_range = f"sp_sindhu_creds!{get_column_letter(id_insert_col)}1:{get_column_letter(id_insert_col)}2"
                        upload_data(1,upload_range,
                                    [[number],[json.dumps(userinfo)]])
                        self.registrationinfo['reg_status'] = 'success'
                        
                    st.divider()
                    st.button("Click to Register",
                              on_click=register_user,
                              args=[str(number)])
        
        else:
            # external user
            st.subheader(":rainbow[Welcome to the SP sindhu registration]")
            st.markdown("### :gray[Just enter few details]")
            with st.popover("Why are we asking phone number"):
                st.markdown("1\. To protect against robotic attacks")
                st.markdown("2\. To have a unique id of user")
                st.markdown("2\. To connect with you if any problems 🤝")
            
            number = st.number_input("Please enter 10 digit phone number",step=1)
            number_verified = False
            if number:
                if len(str(number)) !=10:
                    st.error("Please enter 10 digit number")
                elif not all([i in '0123456789' for i in str(number)]):
                    st.error("Only Numbers are allowed")
                
                elif str(number) in self.userdb['dict']['existing_userid_list']:
                    st.error("A user already exists with this number")
                else:
                    # success
                    st.caption("Jai Haribol!!")
                    number_verified = True
            if  number_verified:
                # move ahead only if number is verified
                
                name = st.text_input("Enter name (without any prefix like ys or your servant )",
                                    max_chars=20)
                password = st.text_input("Enter password",type='password',
                                        max_chars=15)
                if not name:
                    st.caption("Must write name")
                elif not password:
                    st.caption("Must have a password")
                else:
                    # verified
                    def register_user_ext(number,userinfo):
                            password = userinfo['password']
                            name = f"{userinfo['name']} Pr"
                            userinfo = {
                                        "account_is_active":'no',
                                        "creds":{'name':name,
                                                'password':password,
                                                'phone':number}
                                        }
                            id_insert_col = int(self.userdb['dict']['insert_id_col_index'])
                            upload_range = f"sp_sindhu_creds!{get_column_letter(id_insert_col)}1:{get_column_letter(id_insert_col)}2"
                            upload_data(1,upload_range,
                                        [[number],[json.dumps(userinfo)]])
                            self.registrationinfo['reg_status'] = 'success'
                            
                    st.divider()
                    st.button("Click to Register",
                            on_click=register_user_ext,
                            args=[number,{"name":name,
                                   "password":password,
                                   'phone':number}])
        
        st.button("Continue as guest for now",
                      on_click=self.switch_page,
                      args=['SP'])
        if self.bdvapp.userinfo and 'sp_vani_admin' in self.bdvapp.userinfo['roles']:
            # show the pending registrations
            def activate_account(update_range,update_dict):
                update_dict['account_is_active'] = 'yes'
                update_json = json.dumps(update_dict)
                upload_data(1,update_range,[[update_json]])
                
                self._sp_userdb_refresh = True
            
            alldf = self.userdb['dfall']
            for one_user in self.userdb['dict']['existing_userid_list']:
                userdict = json.loads(alldf.query("spid=='userinfo'")[one_user].tolist()[0])
                spdb_col = json.loads(alldf.query("spid=='dbcol'")[one_user].tolist()[0])
                spdb_row = 2
                spdb_update_range = f"sp_sindhu_creds!{get_column_letter(spdb_col)}{spdb_row}"
                if userdict['account_is_active'] == 'no':
                    st.button(f"Activate `{userdict['creds']['name']}`",
                              on_click=activate_account,
                              args=[spdb_update_range,userdict],
                              key="activation_"+one_user)
                    st.divider()
            
    
    def run(self):
        st.markdown(
        """
        <style>
        .step-up,
        .step-down {
            display: none;
        }
        </style>
        </style>
        """,
        unsafe_allow_html=True
        )
        
        if self.userinfo['mode'] == 'guest' and self.current_page !='user_registration':
            # option to login
            with st.sidebar:
                st.info("You are not logged in")
                st.markdown(":gray[Login to have a personalized experience]")
                    
                username = str(st.number_input("Enter username (10digit phone)",min_value=0,step=1))
                password = st.text_input("Enter password",type='password')
                if username:
                    login_response = self.perform_login(username,password,'ask')
                    if login_response==0:
                        st.caption("no user exists")
                    elif login_response == 1:
                        st.caption("incorrect password")
                    elif login_response==2:
                        st.button("login",on_click=self.perform_login,
                                  key='sidebar_login_',
                                  args=[username,password,'submit'])
                st.divider()
                st.button("Click Here to register",type='primary',on_click=self.switch_page,
                        args=['user_registration'])
                
        
        self.page_map[self.current_page]()

class VANI_hearing_class:
    
    def __init__(self):
        self.page_config = {'page_title': "VANI Syllabus",
                            'page_icon':'💌',
                            'layout':'centered'}
        self.page_map = {
            'active':'dash',
            'dash':self.dash,
            'hearnow':self.hear_now,
            'leaderboard':self.leaderboard,
            }
        self.mega = Mega()
                
        # Vani syllabus data
        vani_syllabus_df = pd.read_csv("./local_data/vani_syllabus_config.csv")
        vani_syllabus_df.insert(0,'select',False)
        num_cols = ['sp_index','canto','chapter']
        vani_syllabus_df[num_cols] = vani_syllabus_df[num_cols].fillna(int(0))
        num_cols = ['vani_index',*num_cols]
        vani_syllabus_df[num_cols] = vani_syllabus_df[num_cols].astype('int')
        self.vani_syllabus_df = vani_syllabus_df
        
        # user data
        self._userdb = None
        self._userdb_refresh = True
        
        # page data
        self._series_display_summary = None
        self.user_selections = {'speaker':'blank',
                                'semyear':None
                                }
        
        # hear now
        self.page_hearnow = {
            'status':'',
            'id':'',
            'lecture_info_dict':{}
        }
        
    @property
    def userdb(self):
        """
        * config (appended with user_df)
        * alluser_df, center_df, user_df, mdict
        """
        if self._userdb_refresh:
            raw_array = download_sheet(1,'vani_syllabus_creds')
            raw_df = pd.DataFrame(raw_array[1:],columns=raw_array[0])
            
            key_valuedf = raw_df[['mdata_key','mdata_value']].copy(deep=True)
            key_valuedf = raw_df.query("mdata_key!=''").copy()
            alluser_df = raw_df.drop(columns=['mdata_key','mdata_value'])
            
            # process metadata dictionary
            key_valuedict = {k:v for k,v in dict(zip(key_valuedf.mdata_key,
                              key_valuedf.mdata_value)).items() if k[0]!='_'}
            key_valuedict['dbcol_dict'] = json.loads(key_valuedict['dbcol_dict'])
            
            # process alluser_datadf 
            cols = ['dbrow','vani_id']
            alluser_df[cols] = alluser_df[cols].astype('int')
            alluser_df = alluser_df.query("vani_id!=0").copy()
            
            # slice the data for the current user
            def slice_userinfo(user_column_id):
                columns = ['dbrow','vani_id',user_column_id]
                user_df = alluser_df[columns].copy(deep=True)
                user_df.columns = ['dbrow','vani_id','info']
                user_df['info'] = user_df['info'].apply(lambda x: {'status':'pending'} if x=='' else json.loads(x))
                user_df['status'] = user_df['info'].apply(lambda x: x['status']) # pending, in_progress, done
                return user_df
            
            current_userinfo = self.bdv.userinfo
            user_df = slice_userinfo(f"{current_userinfo['center_name']}_{current_userinfo['phone_number']}")
            config_appended = pd.merge(self.vani_syllabus_df,user_df,
                                       how='left',left_on='vani_index',right_on='vani_id')
            # to verify the merge
            # st.write(user_df.shape)
            # st.write(self.vani_syllabus_df.shape)
            # st.write(config_appended.shape)
            
            
            self._userdb = {
                # alluser_df, center_df, user_df, mdict
                            'config':config_appended,
                            'alluser_df':alluser_df,
                            # 'center_df':,
                            'user_df':user_df,
                            'mdict':key_valuedict
                            }
            self._userdb_refresh = False
        
        return self._userdb
    
    @property
    def bdv(self):
        return st.session_state['bdv_app']
    
    def __single_select_chb(self,editor_key,checkbox_column):
        """
        * original_key = the df should be stored in st.session_state[original_key]
        * checkbox_column = name of the checkbox column
        * It must be the first column"""
        edited_rows = st.session_state[editor_key]['edited_rows']
        if len(edited_rows)==1:
            row_num = list(edited_rows.keys())[0]
            # df = st.session_state[original_key].copy(deep=True)
            df = self._series_display_summary.copy()
            
            set_2_True = list(edited_rows.values())[0][checkbox_column]        
            if set_2_True:
                df[checkbox_column] = False
                df.iloc[row_num,0] = True
                self._series_display_summary= df.copy()
            else:
                df[checkbox_column] = False
                self._series_display_summary= df.copy()
                st.session_state[editor_key]['edited_rows'] = {}
        
        
    def dash(self):
        st.header(":rainbow[Vaani Syllabus @BDV]")
        
        
        # access management
        if self.bdv.userinfo['vani_syllabus_status'] == '':
            # do not have access
            # give an option to raise request
            def request_access():
                _row = self.bdv.userinfo['db_row']
                upload_data(1,f'creds_v2!S{_row}',[['pending']])
                self.bdv.reset_userdb_creds()
                
            st.button("Request access to VANI Syllabus",
                      on_click=request_access,
                      key="vani ask access")
        elif self.bdv.userinfo['vani_syllabus_status'] == 'pending':
            # display the message that please wait
            st.info("Your request to access is in progress")
            _format_message = '\n'.join([
                'Hare Krishna Pr',
                'I have raised request for access to Vani Syllabus',
                'could you please process it fast',
                'I am very eager to hear the transcendental sound vibrations and perfect my life',
                'ys',
                f"{self.bdv.userinfo['full_name'][:-3]}",
            ])
            st.markdown(f"[message to speed up approval](https://wa.me/917260869161?text={quote_plus(_format_message)})")
            st.markdown("")
            st.info("upon approval you will be able to access")
            
        if 'admin' in self.bdv.userinfo['global_roles']:
            # display options to approve
            login_class = self.bdv.page_map['login']
            user_df = login_class.userdb['user_df']
            pending_df = user_df.query("vani_syllabus_status=='pending'")
            if st.checkbox(f"{len(pending_df)} Pending Requests"):
                st.markdown("You are seeing this because you are an admin")
                def approve_access_to_vani(row_db,user_column_id):
                    upload_data(1,
                                f'creds_v2!S{row_db}',
                                [['active']])
                    next_user_col = int(self.userdb['mdict']['next_user_col'])
                    upload_data(1,
                                f"vani_syllabus_creds!{get_column_letter(next_user_col)}1",
                                [[user_column_id]])
                    self.bdv.reset_userdb_creds()
                
                for _row,row in pending_df.iterrows():
                    st.markdown(f"### :orange[{row['center_name']}] -- :gray[{row['full_name']}]")
                    with st.expander('show details',False):
                        st.write(row.to_dict())
                        st.divider()
                        link_for_vani = f"{st.secrets['prod']['site_url']}?target=vani&mode=user&user={row['full_username']}&pass={row['password']}&keepQuery=yes"
                        _format_message = ''.join(["Hare Krishna ",f"{row['full_name']}\n",
                                                "Your access for VANI syllabus has been approved!!\n",
                                                "You can directly vani page with your credentials using following link\n\n",
                                                f"{link_for_vani}\n\n",
                                                f"Please bookmark this URL for accesing quicly and (or) add to home screen\n",
                                                "Your servant\nshivendra"])
                            
                        st.markdown(f"[notify b4 approving](https://wa.me/91{row['phone_number']}?text={quote_plus(_format_message)})")
                        st.button(f"Approve {row['full_name']}",
                                on_click=approve_access_to_vani,
                                args=[row['db_row'],f"{row['center_name']}_{row['phone_number']}"],
                                )
            
        if self.bdv.userinfo['vani_syllabus_status'] != 'active':
            return
        
        userinfo = self.bdv.userinfo
        st.subheader(f":gray[Hare Krishna ] :blue[{userinfo['full_name']}]")
        
        st.caption("You have following lectures which are in progress")
        # display the lectures which are in progress
        vani_progressdf = self.userdb['config'].query("status == 'in_progress' ").copy()
        if vani_progressdf.shape[0]>0:
            with st.expander(f":orange[{vani_progressdf.shape[0]}] lectures in progress", True):
                base_url = st.secrets['prod']['site_url']
                for _row, row in vani_progressdf.iterrows():
                    # color = {'completed':':green',
                    #         'in_progress':':orange',
                    #         'pending':':red'}[row['status']]
                    color = ':orange'
                    lecture_title = row['display_name']

                    url = f"{base_url}?target=hear_vani&id={row['encrypt_id']}&mode=user&user={userinfo['full_username']}&pass={userinfo['password']}"
                    final_title = f":gray[{lecture_title}] :gray[--] [Resume from {row['info']['heard_until']} mins]({url})"
                    st.markdown(final_title)
                    
        st.divider()
        
        
        st.markdown("")
        st.markdown("")
        speaker_dict = {0:['nak','Nakula Group'],
                        1:['dd','DD Series'],
                        2:['hgrsp','HG RSP'],
                        3:['hhrnsm','HH RNSM'],
                        4:['sp','Srila Prabhupada']}
        
        cols = st.columns(5,gap='small')
        
        def update_series_display_summary(active_speaker):
            if self.user_selections['speaker'] == active_speaker:
                # de-select the active speaker
                self.user_selections.__setitem__('speaker','blank')
                return
                        
            self.user_selections.__setitem__('speaker',active_speaker)
            
            # based on speaker one should get the various series for that speaker
            vanidf = self.userdb['config']
            vanidf = vanidf.query(f"speaker == '{active_speaker}'").copy()
            vanidf['done_flag'] = vanidf['status'].apply(lambda x:1 if x=='completed' else 0)
            vanidf_grouped = vanidf.groupby(by='category').agg({'file_id':'count',
                                                                'done_flag':'sum'
                                                                }).reset_index()
            vanidf_grouped.columns = ['category','total','done']
            vanidf_grouped['display_order'] = vanidf_grouped['category'].apply(lambda x: int(x.split('.')[0]))
            vanidf_grouped['display_category'] = vanidf_grouped['category'].apply(lambda x: x.split('.')[1].strip() )
            
            vanidf_grouped.sort_values(by='display_order',inplace=True)
            vanidf_grouped.drop(columns='display_order', inplace=True)
            vanidf_grouped['select'] = [True,*[False for _ in range(len(vanidf_grouped)-1)]]
            vanidf_grouped.columns = ['category','total','done','display_category','select']
            vanidf_grouped = vanidf_grouped[['select','category','display_category','total','done']]
            
            self._series_display_summary = vanidf_grouped.copy()

        
        for i in range(len(cols)):
            skey,sname = speaker_dict[i]
            is_selected = self.user_selections['speaker'] == skey
            cols[i].button(label=sname,
                           key=f"chosen_speaker_{skey}",
                           on_click=update_series_display_summary,
                           args=[skey],
                           type="primary" if is_selected else "secondary")                        
                
        st.divider()
        if self.user_selections['speaker'] == 'blank':
            st.caption("Please select a speaker to filter results")
            st.info("currently viewing all speakers")
        else:
            # display series only if speaker is selected
            
            column_config = {'select':st.column_config.CheckboxColumn("☑",disabled=False),
                            'display_category':st.column_config.TextColumn("series",disabled=True),
                            'total':st.column_config.NumberColumn("total",disabled=True,help='Total available lectures'),
                            'done':st.column_config.NumberColumn("done", disabled=True, help='Completed lectures')}
            chosen_series_list = st.data_editor(self._series_display_summary,
                                column_config=column_config,
                                column_order=['select','display_category','total','done'],
                                on_change=self.__single_select_chb,
                                args=['series_display_summary','select'],
                                key='series_display_summary',
                                hide_index=True,
                                use_container_width=False).query("select == True")['category'].tolist()
            # st.write(chosen_series_list)
            if not chosen_series_list:
                st.caption("Please Select a Series to filter")
                st.info("currently viewing all series")

        vanidf = self.userdb['config'].copy(deep=True)
        active_speaker = self.user_selections['speaker']
        
        if active_speaker =='blank':
            # no user selected
            # apply no filter
            pass
        elif not chosen_series_list:
            # no series selected
            # apply no filter of series
            vanidf = vanidf.query(f"speaker == '{active_speaker}' ").copy()
        else:
            # both are selected
            # st.write(chosen_series)
            chosen_series = chosen_series_list[0]
            vanidf = vanidf.query(f"speaker == '{active_speaker}' and category == '{chosen_series}'").copy()        
        
        
        
        # to filter out results
        search_query = st.text_input(":gray[Enter query to filter]",max_chars=15).strip()
        st.divider()
        if search_query:
            vanidf = vanidf[vanidf.display_name.str.lower().str.contains(search_query.strip().lower())]
        
        
        vanidf.reset_index(drop=True, inplace=True)
        # st.dataframe(vanidf)

        # display the full lectures
        base_url = st.secrets['prod']['site_url']
        st.caption(f":gray[{len(vanidf)} records found]")
        
        for _row, row in vanidf.iterrows():
            color = {'completed':':green',
                        'in_progress':':orange',
                        'pending':':red'}[row['status']]
            lecture_title = row['display_name']
            
            url = f"{base_url}?target=hear_vani&id={row['encrypt_id']}&mode=user&user={userinfo['full_username']}&pass={userinfo['password']}"
            final_title = f":gray[{_row+1}\.] {color}[{lecture_title}] :gray[--] [click to hear]({url})"
            if search_query:
                # final_title = final_title.replace(search_query,f":orange[{search_query}]")
                if color ==':orange':
                    final_title = re.sub(re.escape(search_query), f":blue[{search_query}]", final_title, flags=re.IGNORECASE)
                else:
                    final_title = re.sub(re.escape(search_query), f":orange[{search_query}]", final_title, flags=re.IGNORECASE)
                    
            st.markdown(final_title)
            
            if (_row+1) %10==0:
                chb_placeholder= st.empty()
                chb_response = chb_placeholder.checkbox(f":blue[Show `{len(vanidf)-1-_row}` more]",
                                    key=f'showmore_{_row}',
                                    value=False)
                if not chb_response:
                    break
                else:
                    chb_placeholder.empty()
       
        st.markdown("")
        st.markdown("")
        st.markdown("")
        st.markdown("")
        st.markdown("")
        st.divider()
        st.subheader(":rainbow[Keep Transcendental knowledge one click away]")
        st.markdown("Just save following url as bookmark or shortcut")
        app_url = st.secrets['prod']['site_url']
        with st.popover("contains password"):
            st.markdown(f"{app_url}?target=vani&mode=user&user={self.bdv.userinfo['full_username']}&pass={self.bdv.userinfo['password']}&keepQuery=yes")
            
    def hear_now(self):
        
        if self.page_hearnow['status'] =='pending':
            # locate the lecture
            lecture_id = self.page_hearnow['id']
            lecturedb = self.userdb['config']
            lecture_df = lecturedb.query(f"encrypt_id == '{lecture_id}'").reset_index(drop=True)
            lecture_dict = lecture_df.to_dict(orient='index')[0]
            
            # save this in the lecture_info_dict
            self.page_hearnow['lecture_info_dict'] = lecture_dict
            self.page_map['active'] = 'hearnow'
            
            self.page_config = {'page_title': lecture_dict['display_name'],
                                    'page_icon':'🎧',
                                    'layout':'centered'}
            
            self.page_map['status'] = 'done'
            
            
        userinfo = self.bdv.userinfo
        lecinfo = self.page_hearnow['lecture_info_dict']

        # st.write(self.page_hearnow['lecture_info_dict'])
        st.markdown(f"## :rainbow[{lecinfo['display_name']}]")
        
        # get the lecture info
        server_type = lecinfo['server_type'] # mega, drive, ytube
        file_id = lecinfo['file_id']

        filename = f"{lecinfo['vani_index']}.mp3"
        
        destination = './local_data/vani_storage'        
        available_file_raw = os.listdir(destination)
        available_files = [i.split("^")[1] for i in available_file_raw]
        
        # if 'admin' in self.bdv.userinfo['global_roles']:
        #     with st.expander("see all available files", expanded=False):
        #         # timestamp = .strftime("%Y%b%d%a %H%M%S")
        #         available_dict = {f:
        #             (datetime.date.fromtimestamp(
        #                 os.path.getmtime(f"{destination}/{f}"))
        #              .strftime("%Y%b%d%a %H%M%S"))
                    
        #             for f in available_files}
        #         st.write(available_dict)
        
        # download only if file not present
        if filename not in available_files:
            # storage management
            MAX_LEC_STORE = 6
            # keep the latest MAX_LEC_STORE files and delete the older ones
            # reverse = True sorted in decending order
            sorted_available_files = sorted(available_file_raw,
                                     key=lambda x: x.split('^')[0],
                                     reverse=True)
            for one_file in sorted_available_files[MAX_LEC_STORE:]:
                # st.write(f"deleting {one_file}")
                os.remove(f"{destination}/{one_file}")
                
        
            msgbox = st.empty()
            download_url = ''
            india_timezone = pytz.timezone('Asia/Kolkata')
            file_prefix = (datetime.datetime.now(india_timezone)
                     .strftime("%d%H%M%S"))
            filename = f"{file_prefix}^{filename}"
            
            if server_type == 'mega':
                download_url = f"https://mega.co.nz/#!{file_id}"
                with msgbox.container():
                    st.markdown(f"downloading from mega ")
                    st.caption("This may take a while please wait ...")
                    self.mega.download_url(download_url,destination,dest_filename=filename)
                    st.success("completed")
                msgbox.empty()
            
            elif server_type =='drive':
                download_url = f"https://drive.google.com/uc?id={file_id}&export=download"
                url = f"https://drive.google.com/uc?id={file_id}"
                with msgbox.container():
                    st.markdown(f"downloading from drive ")
                    st.caption("This may take a while please wait ...")
                    gdown.download(url, f"{destination}/{filename}", quiet=False, fuzzy=True)
                    st.success("done")
                msgbox.empty()
        else:
            # change the filename so that it points to existing file
            existing_file = [i for i in available_file_raw if i.split("^")[1] == filename]
            filename = existing_file[0]
            
            # define the download urls etc
            if server_type =='mega':
                download_url = download_url = f"https://mega.co.nz/#!{file_id}"
            elif server_type =='drive':
                download_url = f"https://drive.google.com/uc?id={file_id}&export=download"
                url = f"https://drive.google.com/file/d/{file_id}/view"
                
        # now display the lecture        
        st.markdown("")
        st.markdown("")
        
        # display the image if the audio file have one
        eye_file = eyed3.load(f"{destination}/{filename}")
        file_duration_secs = eye_file.info.time_secs
        if eye_file.tag and eye_file.tag.images:
            cover_image = Image.open(BytesIO(eye_file.tag.images[0].image_data))
            st.image(cover_image)
            
        st.markdown("")
        st.markdown("")
        seek_mins = 0 if lecinfo['status']!='in_progress' else max(1,int(lecinfo['info']['heard_until'])-1)
        foward_min = st.number_input("forward (in min)",step=1,min_value=0,
                                     value=seek_mins)
        st.markdown("")
        st.markdown("")
        st.markdown("")
        st.audio(f"{destination}/{filename}",format="audio/wav",
                 start_time=foward_min*60)
        st.markdown("")
        st.markdown("")
        if server_type =='mega':
            st.markdown(f"[download from mega]({download_url})")
        elif server_type =='drive':
            st.markdown(f"[download from drive]({download_url})")
            st.markdown("")
            st.markdown(f"[play on drive]({url})")




        # note making etc
        
        def update_lecture_status(status_dict,dbrow,dbcol,dbname):
            status_json = json.dumps(status_dict)
            upload_data(db_id=1,
                        range_name=f"{dbname}!{get_column_letter(dbcol)}{dbrow}",
                        value=[[status_json]])
            self._userdb_refresh = True
        
        st.divider()
                
        st.markdown(f"### :gray[Hare Krishna ] :rainbow[{userinfo['full_name']}]")
        
        hearing_status = lecinfo['status']
        dbrow = lecinfo['dbrow']
        dbcol= self.userdb['mdict']['dbcol_dict'][f"{userinfo['center_name']}_{userinfo['phone_number']}"]
        dbcol = int(dbcol)
        dbsheet = 'vani_syllabus_creds'
        
        
        if hearing_status == 'pending':
            st.error("You have not heard this lecutre")
            
        elif hearing_status == 'in_progress':
            last_modification_time = datetime.datetime.strptime(lecinfo['info']['last_modification_time'],"%Y%b%d%a %H%M%S")\
                .strftime("%d %b %Y @ %H:%M")
            st.warning("### You have heard this lecture until")
            st.markdown(f"### `{lecinfo['info']['heard_until']}` :gray[minutes]")
            st.markdown(f"#### :gray[last heard on] {last_modification_time}")
        else:
            # have heard this lecture
            last_modification_time = datetime.datetime.strptime(lecinfo['info']['last_modification_time'],"%Y%b%d%a %H%M%S")\
                .strftime("%d %b %Y @ %H:%M")
            st.success("### You have heard this lecture on ")
            st.markdown(f"#### {last_modification_time}")
        
        st.divider()
        
        
        
        new_status = st.radio(":gray[Update Status of Lecture]",
                     options=['in progress','completed'],
                     index=1 if hearing_status=='completed' else 0,
                     disabled= True if hearing_status=='completed' else False)
        
        india_timezone = pytz.timezone('Asia/Kolkata')
        timestamp = datetime.datetime.now(india_timezone).strftime("%Y%b%d%a %H%M%S")
            
        if new_status =='in progress':
            duration = st.number_input("Heard until (in minute)",min_value=0,step=1,
                                       value=seek_mins+2)
            
            if duration:
                # st.write(lecinfo)
                lec_notes = st.text_area("Lecture Summary",
                                            value="" if hearing_status=='pending' else lecinfo['info']['lecture_notes'],
                                            max_chars=400)
                n_lines, n_words = len(lec_notes.splitlines()),len(lec_notes.split())
                status_prog = {'status':'in_progress',
                                'heard_until':duration,
                                'lecture_notes':lec_notes,
                                'last_modification_time':timestamp
                                }
                st.caption(f"you have written {n_lines} lines {n_words} words")
                st.button("Update status",on_click=update_lecture_status,
                            args=[status_prog,dbrow,dbcol,dbsheet],
                            type='primary')
                
        elif new_status =='completed':
            MIN_LINE = int(max(2,min(12,(5*file_duration_secs)/(30*60))))
            MIN_WORD = int(max(10,min(80,(20*file_duration_secs)/(30*60))))
            
            lec_notes = st.text_area("Lecture Summary",
                                        help=f"Must write at least {MIN_LINE} lines and minimum {MIN_WORD} words in order to mark as completed",
                                        value="" if hearing_status=='pending' else lecinfo['info']['lecture_notes'],
                                        max_chars=400)
            n_lines, n_words = len(lec_notes.splitlines()),len(lec_notes.split())
            st.caption(f"you have written {n_lines} lines {n_words} words")
            if n_lines>MIN_LINE-1 and n_words>MIN_WORD-1:
                                    
                status_compl = {'status':'completed',
                                'heard_until':-1,
                                'lecture_notes':lec_notes,
                                'last_modification_time':timestamp
                                }
                st.button("Update status",on_click=update_lecture_status,
                            args=[status_compl,dbrow,dbcol,dbsheet],
                            type='primary')
                
                
            else:
                if n_lines <MIN_LINE:
                    st.caption(f"{MIN_LINE-n_lines} more lines are needed")
                if n_words<MIN_WORD:
                    st.caption(f"{MIN_WORD-n_words} more words are needed")
    
    def leaderboard(self):
        pass
        
    def run(self):
        st.markdown(
        """
        <style>
        .step-up,
        .step-down {
            display: none;
        }
        </style>
        </style>
        """,
        unsafe_allow_html=True
        )
        if self.bdv.user_exists:
            self.page_map[self.page_map['active']]()
        else:
            # ask to login
            self.bdv.quick_login()